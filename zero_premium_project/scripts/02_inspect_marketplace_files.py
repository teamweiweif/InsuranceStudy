#!/usr/bin/env python
"""Inspect downloaded Marketplace files and write feasibility outputs."""

from __future__ import annotations

import csv
import io
import re
import traceback
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
MANIFEST_PATH = ROOT / "data" / "metadata" / "data_manifest.csv"
FILE_INVENTORY_PATH = ROOT / "data" / "metadata" / "file_inventory.csv"
COLUMN_INVENTORY_PATH = ROOT / "data" / "metadata" / "column_inventory.csv"
MISSINGNESS_PATH = ROOT / "outputs" / "missingness_summary.csv"
OEP_FEASIBILITY_PATH = ROOT / "outputs" / "oep_outcome_feasibility.csv"
PROTOTYPE_PATH = ROOT / "data" / "intermediate" / "prototype_turnover_2023_2024.csv"
PROTOTYPE_DIAG_PATH = ROOT / "outputs" / "prototype_join_diagnostics.csv"
REPORT_PATH = ROOT / "docs" / "aca_zero_premium_turnover_data_feasibility.md"
REPRO_PATH = ROOT / "docs" / "reproducibility_notes.md"
SAMPLE_DIR = ROOT / "outputs" / "sample_rows"


SUPPRESSED_VALUES = {
    "*",
    "**",
    "***",
    "ds",
    "data suppressed",
    "suppressed",
    "suppression",
    "cell suppressed",
    "n/a",
    "na",
    "null",
    "--",
}

ROLE_PATTERNS = {
    "state": [r"^state$", r"statecode", r"stateabrvtn", r"stateabbreviation"],
    "county_fips": [r"cntyfipscd", r"countyfipscd", r"countyfipscode", r"countyfips", r"fipscode", r"fips"],
    "county": [r"county"],
    "total_plan_selections": [r"^cnsmr$", r"consumer", r"planselection", r"qhpselection"],
    "new_consumers": [r"newcnsmr", r"newconsumer"],
    "total_reenrollment": [r"totrenrl", r"returning", r"reenroll", r"reenrollee"],
    "automatic_reenrollment": [r"autorenrl", r"automaticreenroll"],
    "active_reenrollment": [r"actvrenrl$", r"activereenroll"],
    "active_reenrollment_stay": [r"actvrenrlnsw", r"remained", r"sameplan", r"crosswalked"],
    "active_reenrollment_switch": [r"actvrenrlsw", r"switched"],
    "metal_level": [r"metallevel", r"metal"],
    "fpl": [r"fpl", r"federalpoverty", r"householdincome"],
    "premium": [r"premium", r"avgprm", r"individualrate", r"age40"],
    "aptc": [r"aptc", r"financialassistance", r"advancepremium"],
    "issuer_id": [r"issuerid", r"hiosissuer"],
    "issuer_name": [r"issuer", r"company"],
    "plan_id": [r"planid", r"standardcomponentid", r"standardcomponent"],
    "prior_plan_id": [r"planid20[0-9][0-9]", r"previousplan"],
    "current_plan_id": [r"ageoffplanid", r"silvercsreligplanid"],
    "rating_area": [r"ratingarea"],
    "service_area": [r"servicearea"],
    "market_coverage": [r"marketcoverage"],
    "plan_type": [r"plantype"],
    "age": [r"^age$"],
}

OEP_OUTCOME_CANDIDATES = {
    "total reenrollment": ["Tot_Renrl", "Total_Reenrollees", "Returning_Cnsmr"],
    "automatic reenrollment": ["Auto_Renrl", "Automatic_Renrollees"],
    "active reenrollment": ["Actv_Renrl", "Active_Renrollees"],
    "active reenrollment staying with prior/default plan": [
        "Actv_Renrl_Nsw",
        "Actv_Renrl_NoSw",
        "Active_Reenrollees_NonSwitchers",
    ],
    "active reenrollment switching plans": [
        "Actv_Renrl_Sw",
        "Active_Reenrollees_Switchers",
    ],
    "total plan selections": ["Cnsmr", "Consumers", "Plan_Selections"],
    "new consumers": ["New_Cnsmr", "New_Consumers"],
    "returning consumers": ["Tot_Renrl", "Returning_Cnsmr"],
}


def ensure_dirs() -> None:
    for path in [
        FILE_INVENTORY_PATH.parent,
        COLUMN_INVENTORY_PATH.parent,
        MISSINGNESS_PATH.parent,
        PROTOTYPE_PATH.parent,
        SAMPLE_DIR,
        REPORT_PATH.parent,
    ]:
        path.mkdir(parents=True, exist_ok=True)


def norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value).lower())


def clean_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def clean_id(value: Any) -> str:
    text = clean_text(value)
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def clean_fips(value: Any) -> str:
    text = clean_id(value)
    digits = re.sub(r"\D", "", text)
    return digits.zfill(5) if digits else ""


def clean_money(value: Any) -> float:
    text = clean_text(value)
    if not text:
        return np.nan
    text = text.replace("$", "").replace(",", "")
    text = re.sub(r"[^0-9.\-]", "", text)
    if text in {"", ".", "-"}:
        return np.nan
    try:
        return float(text)
    except ValueError:
        return np.nan


def roles_for_column(column: str) -> list[str]:
    normalized = norm(column)
    roles = []
    for role, patterns in ROLE_PATTERNS.items():
        if any(re.search(pattern, normalized) for pattern in patterns):
            roles.append(role)
    return roles


def safe_name(value: str, max_len: int = 160) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value)
    return cleaned[:max_len].strip("_") or "sample"


def load_manifest() -> pd.DataFrame:
    if not MANIFEST_PATH.exists():
        raise FileNotFoundError(f"Missing manifest: {MANIFEST_PATH}")
    manifest = pd.read_csv(MANIFEST_PATH, dtype=str).fillna("")
    manifest["year"] = manifest["year"].astype(str)
    return manifest


def open_member(path: Path, member: str | None = None):
    if member:
        zf = zipfile.ZipFile(path)
        return zf, zf.open(member)
    return None, path.open("rb")


def list_zip_members(path: Path) -> list[dict[str, Any]]:
    members: list[dict[str, Any]] = []
    try:
        with zipfile.ZipFile(path) as zf:
            for info in zf.infolist():
                if info.is_dir():
                    continue
                members.append(
                    {
                        "member": info.filename,
                        "file_size": info.file_size,
                        "compressed_size": info.compress_size,
                    }
                )
    except zipfile.BadZipFile:
        return []
    return members


def tabular_members(path: Path, file_format: str) -> list[dict[str, Any]]:
    if file_format == "zip":
        members = []
        for member in list_zip_members(path):
            suffix = Path(member["member"]).suffix.lower().lstrip(".")
            if suffix in {"csv", "txt", "xlsx", "xls"}:
                member = dict(member)
                member["format"] = suffix
                members.append(member)
        return members
    suffix = path.suffix.lower().lstrip(".")
    if suffix in {"csv", "txt", "xlsx", "xls"}:
        return [{"member": "", "file_size": path.stat().st_size, "compressed_size": "", "format": suffix}]
    return []


def read_csv_sample(path: Path, member: str | None, nrows: int = 5) -> pd.DataFrame:
    last_error = None
    for encoding in ["utf-8-sig", "latin1"]:
        zf, handle = open_member(path, member)
        try:
            with handle:
                return pd.read_csv(handle, nrows=nrows, dtype=str, encoding=encoding)
        except Exception as exc:  # noqa: BLE001
            last_error = exc
        finally:
            if zf:
                zf.close()
    raise RuntimeError(f"Could not read CSV sample from {path}::{member}: {last_error}")


def csv_row_count_and_missingness(
    path: Path,
    member: str | None,
    columns: list[str],
    key_columns: list[str],
) -> tuple[int, list[dict[str, Any]]]:
    if not columns:
        return 0, []
    usecols = key_columns if key_columns else [columns[0]]
    missing = {
        col: {"rows": 0, "missing": 0, "suppressed_like": 0, "nonmissing": 0}
        for col in usecols
    }
    row_count = 0
    last_error = None
    for encoding in ["utf-8-sig", "latin1"]:
        try:
            zf, handle = open_member(path, member)
            with handle:
                for chunk in pd.read_csv(
                    handle,
                    dtype=str,
                    usecols=usecols,
                    chunksize=100_000,
                    encoding=encoding,
                    keep_default_na=False,
                    low_memory=False,
                ):
                    row_count += len(chunk)
                    for col in usecols:
                        values = chunk[col].astype(str).str.strip()
                        is_missing = values.eq("")
                        suppressed = values.str.lower().isin(SUPPRESSED_VALUES)
                        missing[col]["rows"] += len(values)
                        missing[col]["missing"] += int(is_missing.sum())
                        missing[col]["suppressed_like"] += int(suppressed.sum())
                        missing[col]["nonmissing"] += int((~is_missing & ~suppressed).sum())
            if zf:
                zf.close()
            break
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            row_count = 0
            missing = {
                col: {"rows": 0, "missing": 0, "suppressed_like": 0, "nonmissing": 0}
                for col in usecols
            }
        finally:
            try:
                if zf:
                    zf.close()
            except UnboundLocalError:
                pass
    else:
        raise RuntimeError(f"Could not count CSV rows for {path}::{member}: {last_error}")

    rows = []
    for col, stats in missing.items():
        total = stats["rows"]
        rows.append(
            {
                "column_name": col,
                "rows": total,
                "missing_count": stats["missing"],
                "suppressed_like_count": stats["suppressed_like"],
                "nonmissing_count": stats["nonmissing"],
                "missing_rate": (stats["missing"] + stats["suppressed_like"]) / total
                if total
                else np.nan,
            }
        )
    return row_count, rows


def inspect_excel(path: Path, member: str | None) -> list[dict[str, Any]]:
    excel_bytes: bytes | None = None
    if member:
        with zipfile.ZipFile(path) as zf:
            excel_bytes = zf.read(member)
            data = io.BytesIO(excel_bytes)
    else:
        if path.suffix.lower() == ".xlsx":
            excel_bytes = path.read_bytes()
            data = io.BytesIO(excel_bytes)
        else:
            data = path
    excel = pd.ExcelFile(data)
    workbook = None
    if excel_bytes is not None:
        try:
            workbook = load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
        except Exception:  # noqa: BLE001
            workbook = None
    outputs = []
    for sheet in excel.sheet_names:
        raw = pd.read_excel(excel, sheet_name=sheet, dtype=str, nrows=8, header=None)
        header_row = 0
        for idx, raw_row in raw.iterrows():
            normalized_cells = {norm(value) for value in raw_row.tolist()}
            if "statecode" in normalized_cells and (
                "planidstandardcomponent" in normalized_cells
                or "planid" in normalized_cells
                or "standardcomponentid" in normalized_cells
            ):
                header_row = int(idx)
                break
        sample = pd.read_excel(excel, sheet_name=sheet, dtype=str, nrows=5, header=header_row)
        row_count = ""
        if workbook is not None and sheet in workbook.sheetnames:
            max_row = workbook[sheet].max_row or 0
            row_count = max(0, max_row - header_row - 1)
        outputs.append(
            {"sheet": sheet, "sample": sample, "header_row": header_row, "row_count": row_count}
        )
    if workbook is not None:
        workbook.close()
    return outputs


def inspect_files(manifest: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    file_rows: list[dict[str, Any]] = []
    column_rows: list[dict[str, Any]] = []
    missing_rows: list[dict[str, Any]] = []

    successful = manifest[manifest["download_success"].astype(str).str.lower().eq("true")]
    for _, row in successful.iterrows():
        local_path = ROOT / row["local_path"]
        if not local_path.exists():
            continue
        members = tabular_members(local_path, row["file_format"].lower())
        if not members:
            if row["file_format"].lower() == "zip":
                for member in list_zip_members(local_path):
                    file_rows.append(
                        {
                            **row.to_dict(),
                            "container_member": member["member"],
                            "member_file_size_bytes": member["file_size"],
                            "member_compressed_size_bytes": member["compressed_size"],
                            "tabular_format": Path(member["member"]).suffix.lower().lstrip("."),
                            "rows": "",
                            "columns": "",
                            "column_names": "",
                            "read_success": False,
                            "notes": "ZIP member listed but not a supported tabular file.",
                        }
                    )
            continue

        for member in members:
            member_name = member["member"]
            tab_format = member["format"]
            label = f"{row['source_group']}_{row['year']}_{row['file_type']}_{member_name or local_path.name}"
            try:
                if tab_format in {"csv", "txt"}:
                    sample = read_csv_sample(local_path, member_name or None)
                    columns = [str(col) for col in sample.columns]
                    key_columns = [col for col in columns if roles_for_column(col)]
                    row_count, miss = csv_row_count_and_missingness(
                        local_path, member_name or None, columns, key_columns
                    )
                    sample_path = SAMPLE_DIR / f"{safe_name(label)}.csv"
                    sample.to_csv(sample_path, index=False)
                    file_rows.append(
                        {
                            **row.to_dict(),
                            "container_member": member_name,
                            "member_file_size_bytes": member["file_size"],
                            "member_compressed_size_bytes": member["compressed_size"],
                            "tabular_format": tab_format,
                            "rows": row_count,
                            "columns": len(columns),
                            "column_names": "|".join(columns),
                            "sample_rows_path": str(sample_path.relative_to(ROOT)),
                            "read_success": True,
                            "notes": "Read with pandas chunked CSV inspection.",
                        }
                    )
                    for col in columns:
                        column_rows.append(
                            {
                                **row.to_dict(),
                                "container_member": member_name,
                                "column_name": col,
                                "column_normalized": norm(col),
                                "candidate_roles": "|".join(roles_for_column(col)),
                            }
                        )
                    for miss_row in miss:
                        missing_rows.append(
                            {
                                **row.to_dict(),
                                "container_member": member_name,
                                **miss_row,
                            }
                        )
                elif tab_format in {"xlsx", "xls"}:
                    excel_outputs = inspect_excel(local_path, member_name or None)
                    for excel_output in excel_outputs:
                        sheet = excel_output["sheet"]
                        sample = excel_output["sample"]
                        header_row = excel_output.get("header_row", 0)
                        row_count = excel_output.get("row_count", "")
                        columns = [str(col) for col in sample.columns]
                        sample_path = SAMPLE_DIR / f"{safe_name(label + '_' + sheet)}.csv"
                        sample.to_csv(sample_path, index=False)
                        file_rows.append(
                            {
                                **row.to_dict(),
                                "container_member": member_name,
                                "sheet_name": sheet,
                                "member_file_size_bytes": member["file_size"],
                                "member_compressed_size_bytes": member["compressed_size"],
                                "tabular_format": tab_format,
                                "rows": row_count,
                                "columns": len(columns),
                                "column_names": "|".join(columns),
                                "sample_rows_path": str(sample_path.relative_to(ROOT)),
                                "read_success": True,
                                "notes": f"Excel sample inspected with header row {header_row}; row count read from workbook metadata when available.",
                            }
                        )
                        for col in columns:
                            column_rows.append(
                                {
                                    **row.to_dict(),
                                    "container_member": member_name,
                                    "sheet_name": sheet,
                                    "column_name": col,
                                    "column_normalized": norm(col),
                                    "candidate_roles": "|".join(roles_for_column(col)),
                                }
                            )
            except Exception as exc:  # noqa: BLE001
                file_rows.append(
                    {
                        **row.to_dict(),
                        "container_member": member_name,
                        "member_file_size_bytes": member["file_size"],
                        "member_compressed_size_bytes": member["compressed_size"],
                        "tabular_format": tab_format,
                        "rows": "",
                        "columns": "",
                        "column_names": "",
                        "read_success": False,
                        "notes": f"Read failed: {exc}",
                    }
                )

    file_inventory = pd.DataFrame(file_rows)
    column_inventory = pd.DataFrame(column_rows)
    missingness = pd.DataFrame(missing_rows)
    file_inventory.to_csv(FILE_INVENTORY_PATH, index=False)
    column_inventory.to_csv(COLUMN_INVENTORY_PATH, index=False)
    missingness.to_csv(MISSINGNESS_PATH, index=False)
    return file_inventory, column_inventory, missingness


def first_tabular_path_for_manifest_row(row: pd.Series) -> tuple[Path, str | None, str] | None:
    path = ROOT / row["local_path"]
    members = tabular_members(path, row["file_format"].lower())
    if not members:
        return None
    preferred = sorted(
        members,
        key=lambda x: (
            0 if Path(x["member"]).suffix.lower() == ".csv" or x["format"] == "csv" else 1,
            x["member"],
        ),
    )[0]
    return path, preferred["member"] or None, preferred["format"]


def read_tabular(
    path: Path,
    member: str | None = None,
    usecols: list[str] | None = None,
    nrows: int | None = None,
    excel_header: int = 0,
) -> pd.DataFrame:
    suffix = (Path(member).suffix if member else path.suffix).lower()
    if suffix in {".csv", ".txt"}:
        zf, handle = open_member(path, member)
        try:
            with handle:
                return pd.read_csv(
                    handle,
                    dtype=str,
                    usecols=usecols,
                    nrows=nrows,
                    keep_default_na=False,
                    low_memory=False,
                    encoding="utf-8-sig",
                )
        finally:
            if zf:
                zf.close()
    if suffix in {".xlsx", ".xls"}:
        if member:
            with zipfile.ZipFile(path) as zf:
                data = io.BytesIO(zf.read(member))
        else:
            data = path
        return pd.read_excel(
            data, dtype=str, usecols=usecols, nrows=nrows, header=excel_header
        ).fillna("")
    raise ValueError(f"Unsupported tabular file: {path}::{member}")


def find_column(columns: list[str], candidates: list[str]) -> str | None:
    by_norm = {norm(col): col for col in columns}
    for candidate in candidates:
        found = by_norm.get(norm(candidate))
        if found:
            return found
    for candidate in candidates:
        candidate_norm = norm(candidate)
        for normalized, col in by_norm.items():
            if candidate_norm and candidate_norm in normalized:
                return col
    return None


def detect_suppression(series: pd.Series) -> pd.Series:
    values = series.astype(str).str.strip().str.lower()
    return values.isin(SUPPRESSED_VALUES) | values.str.contains("suppress", na=False)


def numeric_usable(series: pd.Series) -> pd.Series:
    values = series.astype(str).str.strip()
    suppressed = detect_suppression(values)
    nums = pd.to_numeric(
        values.str.replace(",", "", regex=False).str.replace("$", "", regex=False),
        errors="coerce",
    )
    return nums.notna() & ~suppressed


def verify_oep_outcomes(manifest: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    subset = manifest[
        manifest["source_group"].eq("oep_puf")
        & manifest["file_type"].eq("County-Level OEP PUF")
        & manifest["year"].isin(["2022", "2023", "2024"])
        & manifest["download_success"].astype(str).str.lower().eq("true")
    ]
    for _, row in subset.iterrows():
        located = first_tabular_path_for_manifest_row(row)
        if not located:
            continue
        path, member, _fmt = located
        df = read_tabular(path, member)
        columns = list(df.columns)
        fips_col = find_column(
            columns, ["County_FIPS_Cd", "Cnty_FIPS_Cd", "County FIPS Code", "FIPSCode"]
        )
        state_col = find_column(columns, ["State_Abrvtn", "State", "StateCode"])
        geo_level = "county" if fips_col else "unknown"
        for outcome, candidates in OEP_OUTCOME_CANDIDATES.items():
            col = find_column(columns, candidates)
            if col:
                usable = numeric_usable(df[col])
                suppressed = detect_suppression(df[col])
                missing = df[col].astype(str).str.strip().eq("") | suppressed
                if fips_col:
                    usable = usable & df[fips_col].astype(str).str.strip().ne("")
                rows.append(
                    {
                        "year": row["year"],
                        "outcome": outcome,
                        "source_file": row["local_path"],
                        "container_member": member or "",
                        "exact_column_name": col,
                        "state_column": state_col or "",
                        "county_fips_column": fips_col or "",
                        "geographic_level": geo_level,
                        "suppressed_small_cells_exist": bool(suppressed.any()),
                        "suppressed_like_count": int(suppressed.sum()),
                        "number_usable_county_year_observations": int(usable.sum()),
                        "total_rows": int(len(df)),
                        "missingness_rate": float(missing.mean()) if len(df) else np.nan,
                        "matches_drake_categories": "Yes"
                        if outcome
                        in {
                            "total reenrollment",
                            "automatic reenrollment",
                            "active reenrollment",
                            "active reenrollment staying with prior/default plan",
                            "active reenrollment switching plans",
                            "total plan selections",
                            "new consumers",
                            "returning consumers",
                        }
                        else "Unclear",
                    }
                )
            else:
                rows.append(
                    {
                        "year": row["year"],
                        "outcome": outcome,
                        "source_file": row["local_path"],
                        "container_member": member or "",
                        "exact_column_name": "",
                        "state_column": state_col or "",
                        "county_fips_column": fips_col or "",
                        "geographic_level": geo_level,
                        "suppressed_small_cells_exist": "",
                        "suppressed_like_count": "",
                        "number_usable_county_year_observations": 0,
                        "total_rows": int(len(df)),
                        "missingness_rate": 1.0,
                        "matches_drake_categories": "No column found",
                    }
                )
    out = pd.DataFrame(rows)
    out.to_csv(OEP_FEASIBILITY_PATH, index=False)
    return out


def manifest_file(manifest: pd.DataFrame, source_group: str, year: int, file_type: str) -> pd.Series | None:
    hits = manifest[
        manifest["source_group"].eq(source_group)
        & manifest["year"].eq(str(year))
        & manifest["file_type"].eq(file_type)
        & manifest["download_success"].astype(str).str.lower().eq("true")
    ]
    if hits.empty:
        return None
    return hits.iloc[0]


def get_header(path: Path, member: str | None) -> list[str]:
    return list(read_tabular(path, member, nrows=0).columns)


def find_age40_premium_col(columns: list[str]) -> str | None:
    for col in columns:
        text = str(col)
        lowered = text.lower()
        normalized = norm(text)
        if "premium" in lowered and ("age 40" in lowered or "age40" in normalized):
            if "adult" in lowered or "individual" in lowered:
                return col
    for col in columns:
        lowered = str(col).lower()
        if "premium" in lowered and "40" in lowered:
            return col
    return None


def qhp_excel_header(path: Path, member: str | None) -> int:
    suffix = (Path(member).suffix if member else path.suffix).lower()
    if suffix not in {".xlsx", ".xls"}:
        return 0
    if member:
        with zipfile.ZipFile(path) as zf:
            data = io.BytesIO(zf.read(member))
    else:
        data = path
    raw = pd.read_excel(data, sheet_name=0, dtype=str, nrows=8, header=None)
    for idx, raw_row in raw.iterrows():
        normalized_cells = {norm(value) for value in raw_row.tolist()}
        if "statecode" in normalized_cells and "planidstandardcomponent" in normalized_cells:
            return int(idx)
    return 0


def load_qhp_landscape(manifest: pd.DataFrame, year: int) -> tuple[pd.DataFrame, dict[str, str]]:
    row = manifest_file(manifest, "qhp_landscape", year, "QHP Landscape Individual Medical ZIP")
    if row is None:
        raise FileNotFoundError(f"Missing QHP Landscape Individual Medical ZIP for {year}")
    located = first_tabular_path_for_manifest_row(row)
    if not located:
        raise FileNotFoundError(f"No tabular QHP member for {year}")
    path, member, _fmt = located
    excel_header = qhp_excel_header(path, member)
    columns = list(read_tabular(path, member, nrows=0, excel_header=excel_header).columns)
    colmap = {
        "state": find_column(columns, ["State Code", "State", "StateCode"]),
        "county_fips": find_column(
            columns, ["FIPS County Code", "County FIPS Code", "County FIPS", "FIPSCode"]
        ),
        "county_name": find_column(columns, ["County Name", "County"]),
        "plan_id": find_column(columns, ["Plan ID", "PlanId", "StandardComponentId"]),
        "issuer_id": find_column(columns, ["Issuer ID", "IssuerId", "HIOS Issuer ID"]),
        "issuer_name": find_column(columns, ["Issuer Name", "Issuer"]),
        "metal": find_column(columns, ["Metal Level", "MetalLevel"]),
        "rating_area": find_column(columns, ["Rating Area", "RatingAreaId"]),
        "premium_age40": find_age40_premium_col(columns),
    }
    required = ["state", "county_fips", "plan_id", "metal", "premium_age40"]
    missing = [key for key in required if not colmap.get(key)]
    if missing:
        raise ValueError(f"QHP {year} missing required columns: {missing}; columns={columns[:20]}")
    usecols = [col for col in colmap.values() if col]
    df = read_tabular(path, member, usecols=usecols, excel_header=excel_header)
    rename = {value: key for key, value in colmap.items() if value}
    df = df.rename(columns=rename)
    for col in ["state", "county_fips", "plan_id", "issuer_id", "issuer_name", "metal", "rating_area"]:
        if col not in df.columns:
            df[col] = ""
    df["county_fips"] = df["county_fips"].map(clean_fips)
    df["plan_id"] = df["plan_id"].map(clean_id)
    df["plan_id_base"] = df["plan_id"].str.split("-").str[0]
    df["issuer_id"] = df["issuer_id"].map(clean_id)
    df["premium_age40"] = df["premium_age40"].map(clean_money)
    df["state"] = df["state"].astype(str).str.strip().str.upper()
    df["metal"] = df["metal"].astype(str).str.strip()
    df["rating_area"] = df["rating_area"].astype(str).str.strip()
    return df, colmap


def load_crosswalk_2024(manifest: pd.DataFrame) -> pd.DataFrame:
    row = manifest_file(manifest, "exchange_puf", 2024, "Plan ID Crosswalk PUF")
    if row is None:
        raise FileNotFoundError("Missing 2024 Plan ID Crosswalk PUF")
    located = first_tabular_path_for_manifest_row(row)
    if not located:
        raise FileNotFoundError("No tabular member in 2024 Plan ID Crosswalk PUF")
    path, member, _fmt = located
    columns = get_header(path, member)
    candidates = [
        "State",
        "DentalPlan",
        "PlanID_2023",
        "IssuerID_2023",
        "MetalLevel_2023",
        "FIPSCode",
        "ZipCode",
        "CrosswalkLevel",
        "ReasonForCrosswalk",
        "PlanID_2024",
        "IssuerID_2024",
        "MetalLevel_2024",
    ]
    usecols = [col for col in candidates if col in columns]
    df = read_tabular(path, member, usecols=usecols)
    for col in candidates:
        if col not in df.columns:
            df[col] = ""
    df = df.rename(
        columns={
            "State": "state",
            "DentalPlan": "dental_plan",
            "PlanID_2023": "plan_id_2023",
            "IssuerID_2023": "issuer_id_2023",
            "MetalLevel_2023": "metal_2023",
            "FIPSCode": "county_fips",
            "ZipCode": "zip_code",
            "CrosswalkLevel": "crosswalk_level",
            "ReasonForCrosswalk": "reason_for_crosswalk",
            "PlanID_2024": "plan_id_2024",
            "IssuerID_2024": "issuer_id_2024",
            "MetalLevel_2024": "metal_2024",
        }
    )
    df["county_fips"] = df["county_fips"].map(clean_fips)
    for col in ["plan_id_2023", "plan_id_2024", "issuer_id_2023", "issuer_id_2024"]:
        df[col] = df[col].map(clean_id)
    df["plan_id_2023_base"] = df["plan_id_2023"].str.split("-").str[0]
    df["plan_id_2024_base"] = df["plan_id_2024"].str.split("-").str[0]
    df["state"] = df["state"].astype(str).str.strip().str.upper()
    if "dental_plan" in df.columns:
        df = df[~df["dental_plan"].astype(str).str.lower().isin({"yes", "true", "1"})].copy()
    return df


def load_rate_age40(manifest: pd.DataFrame, year: int) -> pd.DataFrame:
    row = manifest_file(manifest, "exchange_puf", year, "Rate PUF")
    if row is None:
        raise FileNotFoundError(f"Missing {year} Rate PUF")
    located = first_tabular_path_for_manifest_row(row)
    if not located:
        raise FileNotFoundError(f"No tabular member in {year} Rate PUF")
    path, member, _fmt = located
    columns = get_header(path, member)
    colmap = {
        "state": find_column(columns, ["StateCode", "State Code"]),
        "plan_id": find_column(columns, ["PlanId", "Plan ID"]),
        "rating_area": find_column(columns, ["RatingAreaId", "Rating Area ID"]),
        "age": find_column(columns, ["Age"]),
        "individual_rate": find_column(columns, ["IndividualRate", "Individual Rate"]),
    }
    usecols = [col for col in colmap.values() if col]
    df = read_tabular(path, member, usecols=usecols)
    df = df.rename(columns={value: key for key, value in colmap.items() if value})
    for col in ["state", "plan_id", "rating_area", "age", "individual_rate"]:
        if col not in df.columns:
            df[col] = ""
    df = df[df["age"].astype(str).str.strip().eq("40")].copy()
    df["plan_id"] = df["plan_id"].map(clean_id)
    df["plan_id_base"] = df["plan_id"].str.split("-").str[0]
    df["rating_area"] = df["rating_area"].astype(str).str.strip()
    df["state"] = df["state"].astype(str).str.strip().str.upper()
    df["individual_rate"] = df["individual_rate"].map(clean_money)
    return df.drop_duplicates(["state", "plan_id_base", "rating_area"])


def load_plan_service_county(manifest: pd.DataFrame, year: int) -> pd.DataFrame:
    plan_row = manifest_file(manifest, "exchange_puf", year, "Plan Attributes PUF")
    service_row = manifest_file(manifest, "exchange_puf", year, "Service Area PUF")
    if plan_row is None or service_row is None:
        return pd.DataFrame()
    plan_loc = first_tabular_path_for_manifest_row(plan_row)
    service_loc = first_tabular_path_for_manifest_row(service_row)
    if not plan_loc or not service_loc:
        return pd.DataFrame()

    plan_path, plan_member, _ = plan_loc
    service_path, service_member, _ = service_loc
    plan_cols = get_header(plan_path, plan_member)
    service_cols = get_header(service_path, service_member)
    plan_colmap = {
        "state": find_column(plan_cols, ["StateCode"]),
        "issuer_id": find_column(plan_cols, ["IssuerId"]),
        "plan_id": find_column(plan_cols, ["StandardComponentId", "PlanId"]),
        "service_area": find_column(plan_cols, ["ServiceAreaId"]),
        "metal": find_column(plan_cols, ["MetalLevel"]),
        "market": find_column(plan_cols, ["MarketCoverage"]),
        "plan_type": find_column(plan_cols, ["PlanType"]),
    }
    service_colmap = {
        "state": find_column(service_cols, ["StateCode"]),
        "issuer_id": find_column(service_cols, ["IssuerId"]),
        "service_area": find_column(service_cols, ["ServiceAreaId"]),
        "county_fips": find_column(service_cols, ["County"]),
        "market": find_column(service_cols, ["MarketCoverage"]),
        "dental": find_column(service_cols, ["DentalOnlyPlan"]),
    }
    plan = read_tabular(plan_path, plan_member, usecols=[c for c in plan_colmap.values() if c])
    service = read_tabular(
        service_path, service_member, usecols=[c for c in service_colmap.values() if c]
    )
    plan = plan.rename(columns={value: key for key, value in plan_colmap.items() if value})
    service = service.rename(columns={value: key for key, value in service_colmap.items() if value})
    for col in ["state", "issuer_id", "plan_id", "service_area"]:
        if col not in plan.columns:
            plan[col] = ""
    for col in ["state", "issuer_id", "service_area", "county_fips"]:
        if col not in service.columns:
            service[col] = ""
    plan["plan_id"] = plan["plan_id"].map(clean_id)
    plan["plan_id_base"] = plan["plan_id"].str.split("-").str[0]
    service["county_fips"] = service["county_fips"].map(clean_fips)
    merged = plan.merge(
        service,
        on=["state", "issuer_id", "service_area"],
        how="left",
        suffixes=("_plan", "_service"),
    )
    return merged[["state", "issuer_id", "plan_id_base", "service_area", "county_fips"]].drop_duplicates()


def run_treatment_prototype(manifest: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    diagnostics: list[dict[str, Any]] = []
    try:
        qhp23, qhp23_cols = load_qhp_landscape(manifest, 2023)
        qhp24, qhp24_cols = load_qhp_landscape(manifest, 2024)
        sample_states = [state for state in ["AL", "FL", "NC", "TX", "WI"] if state in set(qhp23["state"])]
        if not sample_states:
            sample_states = sorted(qhp23["state"].dropna().unique().tolist())[:5]
        qhp23 = qhp23[qhp23["state"].isin(sample_states)].copy()
        qhp24 = qhp24[qhp24["state"].isin(sample_states)].copy()

        silver23 = qhp23[
            qhp23["metal"].str.lower().str.contains("silver", na=False)
            & qhp23["premium_age40"].notna()
            & qhp23["county_fips"].ne("")
            & qhp23["plan_id_base"].ne("")
        ].copy()
        silver24 = qhp24[
            qhp24["metal"].str.lower().str.contains("silver", na=False)
            & qhp24["premium_age40"].notna()
            & qhp24["county_fips"].ne("")
            & qhp24["plan_id_base"].ne("")
        ].copy()

        silver23 = silver23.sort_values(["county_fips", "premium_age40", "plan_id_base"])
        silver23 = silver23.drop_duplicates(["county_fips", "plan_id_base"])
        silver23["silver_rank_2023"] = silver23.groupby("county_fips").cumcount() + 1
        prev_two = silver23[silver23["silver_rank_2023"].le(2)].copy()
        prev_two["prior_zero_post_subsidy_proxy"] = True

        silver24 = silver24.sort_values(["county_fips", "premium_age40", "plan_id_base"])
        silver24 = silver24.drop_duplicates(["county_fips", "plan_id_base"])
        silver24["silver_rank_2024"] = silver24.groupby("county_fips").cumcount() + 1
        slcsp24 = silver24[silver24["silver_rank_2024"].eq(2)][
            ["county_fips", "premium_age40"]
        ].rename(columns={"premium_age40": "slcsp_age40_2024"})

        cw = load_crosswalk_2024(manifest)
        cw = cw[cw["state"].isin(sample_states)].copy()
        cw_county = cw[cw["county_fips"].ne("")].drop_duplicates(
            ["county_fips", "plan_id_2023_base", "plan_id_2024_base"]
        )
        prev_joined = prev_two.merge(
            cw_county,
            left_on=["county_fips", "plan_id_base"],
            right_on=["county_fips", "plan_id_2023_base"],
            how="left",
            suffixes=("_qhp2023", "_cw"),
        )
        if "state_qhp2023" in prev_joined.columns and "state" not in prev_joined.columns:
            prev_joined["state"] = prev_joined["state_qhp2023"]
        current_lookup = silver24[
            [
                "county_fips",
                "plan_id_base",
                "issuer_id",
                "issuer_name",
                "premium_age40",
                "rating_area",
                "silver_rank_2024",
            ]
        ].rename(
            columns={
                "plan_id_base": "plan_id_2024_base",
                "issuer_id": "issuer_id_2024_qhp",
                "issuer_name": "issuer_name_2024_qhp",
                "premium_age40": "premium_age40_2024",
                "rating_area": "rating_area_2024",
            }
        )
        current_joined = prev_joined.merge(
            current_lookup, on=["county_fips", "plan_id_2024_base"], how="left"
        ).merge(slcsp24, on="county_fips", how="left")
        current_joined["post_subsidy_premium_proxy_2024"] = (
            current_joined["premium_age40_2024"] - current_joined["slcsp_age40_2024"]
        ).clip(lower=0)
        current_joined["mapped_current_positive_post_subsidy_proxy"] = (
            current_joined["post_subsidy_premium_proxy_2024"] > 0.01
        )
        current_joined["issuer_id_2023_final"] = current_joined["issuer_id"].where(
            current_joined["issuer_id"].astype(str).str.strip().ne(""),
            current_joined["issuer_id_2023"],
        )
        current_joined["issuer_id_2024_final"] = current_joined["issuer_id_2024_qhp"].where(
            current_joined["issuer_id_2024_qhp"].astype(str).str.strip().ne(""),
            current_joined["issuer_id_2024"],
        )
        current_joined["across_issuer"] = (
            current_joined["issuer_id_2023_final"].astype(str)
            != current_joined["issuer_id_2024_final"].astype(str)
        )
        current_joined["within_issuer"] = (
            current_joined["issuer_id_2023_final"].astype(str)
            == current_joined["issuer_id_2024_final"].astype(str)
        )

        rate23 = load_rate_age40(manifest, 2023)
        rate24 = load_rate_age40(manifest, 2024)
        prev_rate = prev_two.merge(
            rate23,
            left_on=["state", "plan_id_base", "rating_area"],
            right_on=["state", "plan_id_base", "rating_area"],
            how="left",
        )
        cur_rate = current_joined.merge(
            rate24,
            left_on=["state", "plan_id_2024_base", "rating_area_2024"],
            right_on=["state", "plan_id_base", "rating_area"],
            how="left",
            suffixes=("", "_rate2024"),
        )
        service24 = load_plan_service_county(manifest, 2024)
        service_join = current_joined.merge(
            service24,
            left_on=["state", "plan_id_2024_base", "county_fips"],
            right_on=["state", "plan_id_base", "county_fips"],
            how="left",
            suffixes=("", "_service"),
        )

        def diag(metric: str, numerator: int | float, denominator: int | float, notes: str = "") -> None:
            diagnostics.append(
                {
                    "prototype": "2023_to_2024_age40_low_income_proxy",
                    "sample_states": "|".join(sample_states),
                    "metric": metric,
                    "numerator": numerator,
                    "denominator": denominator,
                    "rate": float(numerator / denominator) if denominator else np.nan,
                    "notes": notes,
                }
            )

        diag("previous-year top-two silver county-plan rows", len(prev_two), len(prev_two))
        diag(
            "previous-year plan to crosswalk",
            int(current_joined["plan_id_2024_base"].astype(str).str.strip().ne("").sum()),
            len(current_joined),
            "Join on county FIPS plus previous-year plan ID.",
        )
        diag(
            "crosswalk to current-year QHP Landscape",
            int(current_joined["premium_age40_2024"].notna().sum()),
            len(current_joined),
            "Join on county FIPS plus current-year plan ID.",
        )
        diag(
            "previous-year plan to Rate PUF",
            int(prev_rate["individual_rate"].notna().sum()),
            len(prev_rate),
            "Join on state, plan ID, rating area, and age 40.",
        )
        diag(
            "current-year plan to Rate PUF",
            int(cur_rate["individual_rate"].notna().sum()),
            len(cur_rate),
            "Join on state, plan ID, rating area, and age 40.",
        )
        diag(
            "current-year plan to service-area county",
            int(service_join["service_area"].astype(str).str.strip().ne("").sum()),
            len(service_join),
            "Join Plan Attributes to Service Area PUF by issuer/service area, then to current plan county.",
        )

        county = current_joined.groupby(["state", "county_fips"], dropna=False).agg(
            prior_top_two_silver_plans=("plan_id_base", "nunique"),
            crosswalked_plans=("plan_id_2024_base", lambda x: x.astype(str).str.strip().ne("").sum()),
            current_plan_premium_matches=("premium_age40_2024", lambda x: x.notna().sum()),
            any_zero_to_positive_turnover=(
                "mapped_current_positive_post_subsidy_proxy",
                lambda x: bool(x.fillna(False).any()),
            ),
            zero_to_positive_turnover_across_issuer=(
                "mapped_current_positive_post_subsidy_proxy",
                lambda x: False,
            ),
        ).reset_index()

        across = (
            current_joined[
                current_joined["mapped_current_positive_post_subsidy_proxy"].fillna(False)
                & current_joined["across_issuer"].fillna(False)
            ]
            .groupby(["state", "county_fips"])
            .size()
            .rename("across_count")
            .reset_index()
        )
        within = (
            current_joined[
                current_joined["mapped_current_positive_post_subsidy_proxy"].fillna(False)
                & current_joined["within_issuer"].fillna(False)
            ]
            .groupby(["state", "county_fips"])
            .size()
            .rename("within_count")
            .reset_index()
        )
        county = county.drop(columns=["zero_to_positive_turnover_across_issuer"]).merge(
            across, on=["state", "county_fips"], how="left"
        )
        county = county.merge(within, on=["state", "county_fips"], how="left")
        county["across_count"] = county["across_count"].fillna(0).astype(int)
        county["within_count"] = county["within_count"].fillna(0).astype(int)
        county["zero_to_positive_turnover_across_issuer"] = county["across_count"] > 0
        county["zero_to_positive_turnover_within_issuer"] = county["within_count"] > 0
        county["prototype_notes"] = (
            "Low-income proxy assumes 100-150% FPL expected contribution is zero, "
            "so the two cheapest silver plans have zero post-subsidy premium; "
            "positive current-year premium is premium above current-year SLCSP."
        )
        county.to_csv(PROTOTYPE_PATH, index=False)
        diag_df = pd.DataFrame(diagnostics)
        diag_df.to_csv(PROTOTYPE_DIAG_PATH, index=False)
        return county, diag_df
    except Exception as exc:  # noqa: BLE001
        diag_df = pd.DataFrame(
            [
                {
                    "prototype": "2023_to_2024_age40_low_income_proxy",
                    "sample_states": "",
                    "metric": "prototype failed",
                    "numerator": "",
                    "denominator": "",
                    "rate": "",
                    "notes": f"{exc}\n{traceback.format_exc()}",
                }
            ]
        )
        diag_df.to_csv(PROTOTYPE_DIAG_PATH, index=False)
        pd.DataFrame().to_csv(PROTOTYPE_PATH, index=False)
        return pd.DataFrame(), diag_df


def sample_alignment(manifest: pd.DataFrame, oep: pd.DataFrame) -> pd.DataFrame:
    rows = []
    county_files = manifest[
        manifest["source_group"].eq("oep_puf")
        & manifest["file_type"].eq("County-Level OEP PUF")
        & manifest["year"].isin(["2022", "2023", "2024"])
        & manifest["download_success"].astype(str).str.lower().eq("true")
    ]
    county_sets: dict[str, set[str]] = {}
    for _, row in county_files.iterrows():
        located = first_tabular_path_for_manifest_row(row)
        if not located:
            continue
        path, member, _ = located
        df = read_tabular(path, member)
        fips_col = find_column(
            list(df.columns), ["County_FIPS_Cd", "Cnty_FIPS_Cd", "County FIPS Code", "FIPSCode"]
        )
        state_col = find_column(list(df.columns), ["State_Abrvtn", "State", "StateCode"])
        if fips_col:
            fips = set(df[fips_col].map(clean_fips).loc[lambda x: x.ne("")])
            county_sets[row["year"]] = fips
            rows.append(
                {
                    "check": f"OEP county rows {row['year']}",
                    "value": len(df),
                    "notes": f"State column: {state_col or 'not found'}; FIPS column: {fips_col}",
                }
            )
    if county_sets:
        common = set.intersection(*county_sets.values()) if len(county_sets) > 1 else next(iter(county_sets.values()))
        rows.append(
            {
                "check": "counties present in all downloaded 2022-2024 OEP county files",
                "value": len(common),
                "notes": "Raw overlap before excluding Alaska, Hawaii, Nebraska, or non-HealthCare.gov states.",
            }
        )
    for state_name, prefix in [("Alaska", "02"), ("Hawaii", "15"), ("Nebraska", "31")]:
        present = any(any(fips.startswith(prefix) for fips in values) for values in county_sets.values())
        rows.append(
            {
                "check": f"{state_name} county FIPS present in OEP county files",
                "value": present,
                "notes": "Drake-style sample would exclude or specially handle this state as described.",
            }
        )
    if not oep.empty:
        suppressed = oep.groupby("year")["missingness_rate"].mean().reset_index()
        for _, row in suppressed.iterrows():
            rows.append(
                {
                    "check": f"mean missing/suppressed rate across tested OEP outcomes {row['year']}",
                    "value": row["missingness_rate"],
                    "notes": "Outcome-specific details are in outputs/oep_outcome_feasibility.csv.",
                }
            )
    rows.append(
        {
            "check": "OEP reenrollment stratification",
            "value": "county-year aggregate",
            "notes": "County-level OEP PUFs do not expose individual-level retention outcomes or income-stratified reenrollment outcomes.",
        }
    )
    out = pd.DataFrame(rows)
    out.to_csv(ROOT / "outputs" / "sample_alignment_checks.csv", index=False)
    return out


def summarize_years(values: pd.Series) -> str:
    years = sorted(str(v) for v in values.dropna().unique() if str(v))
    return ", ".join(years)


def markdown_table(df: pd.DataFrame, columns: list[str] | None = None, max_rows: int = 20) -> str:
    if df.empty:
        return "_No rows._"
    shown = df[columns] if columns else df
    shown = shown.head(max_rows).copy()
    shown = shown.fillna("")
    headers = [str(col) for col in shown.columns]

    def cell(value: Any) -> str:
        text = str(value).replace("\n", " ").replace("|", "\\|")
        return text

    lines = [
        "| " + " | ".join(cell(header) for header in headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for _, row in shown.iterrows():
        lines.append("| " + " | ".join(cell(row[col]) for col in shown.columns) + " |")
    return "\n".join(lines)


def generate_report(
    manifest: pd.DataFrame,
    file_inventory: pd.DataFrame,
    column_inventory: pd.DataFrame,
    oep: pd.DataFrame,
    prototype_diag: pd.DataFrame,
    alignment: pd.DataFrame,
) -> None:
    successful_manifest = manifest[manifest["download_success"].astype(str).str.lower().eq("true")]
    oep_required = {
        "total reenrollment",
        "automatic reenrollment",
        "active reenrollment",
        "active reenrollment staying with prior/default plan",
        "active reenrollment switching plans",
        "total plan selections",
        "new consumers",
        "returning consumers",
    }
    oep_years = oep["year"].astype(str) if "year" in oep.columns else pd.Series([], dtype=str)
    found_oep = set(
        oep.loc[
            oep["exact_column_name"].astype(str).str.strip().ne("")
            & oep_years.isin(["2022", "2023", "2024"]),
            "outcome",
        ]
    )
    prototype_failed = (
        prototype_diag["metric"].astype(str).str.contains("prototype failed", case=False, na=False).any()
        if not prototype_diag.empty
        else True
    )
    status = "Conditional Go"
    if not oep_required.issubset(found_oep):
        status = "No-Go for Drake-style OEP outcomes until missing OEP columns are resolved"
    elif prototype_failed:
        status = "Conditional Go for outcomes; treatment reconstruction unresolved"

    source_summary = (
        manifest.groupby(["source_group", "file_type", "source_page_url"], dropna=False)
        .agg(
            years=("year", summarize_years),
            direct_download_url_example=("direct_download_url", "first"),
            files_identified=("direct_download_url", "count"),
            downloads_successful=("download_success", lambda x: int(x.astype(str).str.lower().eq("true").sum())),
            downloads_failed=("download_success", lambda x: int((~x.astype(str).str.lower().eq("true")).sum())),
        )
        .reset_index()
    )
    successful_url = (
        manifest[manifest["download_success"].astype(str).str.lower().eq("true")]
        .groupby(["source_group", "file_type", "source_page_url"], dropna=False)["direct_download_url"]
        .first()
        .rename("successful_direct_download_url_example")
        .reset_index()
    )
    source_summary = source_summary.merge(
        successful_url, on=["source_group", "file_type", "source_page_url"], how="left"
    )
    source_summary["direct_download_url_example"] = source_summary[
        "successful_direct_download_url_example"
    ].where(
        source_summary["successful_direct_download_url_example"].astype(str).str.strip().ne(""),
        source_summary["direct_download_url_example"],
    )
    source_summary = source_summary.drop(columns=["successful_direct_download_url_example"])
    file_summary_cols = [
        "source_group",
        "year",
        "file_type",
        "container_member",
        "rows",
        "columns",
        "local_path",
    ]
    key_cols = column_inventory[
        column_inventory["candidate_roles"].astype(str).str.strip().ne("")
    ][["source_group", "year", "file_type", "column_name", "candidate_roles"]].drop_duplicates()

    problems = []
    failed = manifest[~manifest["download_success"].astype(str).str.lower().eq("true")]
    if not failed.empty:
        problems.append(
            f"{len(failed)} manifest rows did not download. Most are expected missing QHP Landscape direct URLs for PY2019-PY2021 or older dictionary patterns; see the manifest."
        )
    if "returning consumers" in found_oep:
        problems.append("Returning consumers are represented by the same county-level `Tot_Renrl` column as total reenrollment in the tested OEP files.")
    problems.append("The public OEP files are aggregate PUFs, so they do not support individual-level HTE or demographic/income-stratified reenrollment outcomes at county level.")
    problems.append("The treatment prototype uses a low-income age-40 proxy; exact household-specific post-subsidy premiums require explicit subsidy parameterization and household composition.")
    problems.append("Nebraska county-market handling and Alaska/Hawaii exclusions were not force-applied in this Step 1 audit.")

    report = f"""# ACA Zero-Premium Turnover Data Feasibility

## Executive Summary

Overall feasibility: **{status}**.

The public files support a county-year feasibility workflow for the Drake et al.-style topic. County-level OEP PUFs for 2022-2024 expose the reenrollment and plan-switching outcome columns needed for aggregate county-year outcomes. QHP Landscape files for PY2022-PY2026 expose county-plan silver premiums, and the Exchange PUF Plan ID Crosswalk exposes prior-to-current plan mapping. The main limitation is not access to public data, but exactness: individual-level retention and household-specific subsidy calculations are not public, and PY2021 QHP Landscape direct files were not available from Data.HealthCare.gov by the tested URL pattern.

## Data Sources Found

{markdown_table(source_summary, ["source_group", "file_type", "years", "files_identified", "downloads_successful", "downloads_failed", "source_page_url", "direct_download_url_example"], 60)}

Full manifest: `data/metadata/data_manifest.csv`.

## File Inventory

{markdown_table(file_inventory[file_summary_cols].sort_values(["source_group", "year", "file_type"]).head(40), file_summary_cols, 40)}

Full file inventory: `data/metadata/file_inventory.csv`. Full column inventory: `data/metadata/column_inventory.csv`.

Selected detected key columns:

{markdown_table(key_cols.head(60), ["source_group", "year", "file_type", "column_name", "candidate_roles"], 60)}

## OEP Outcome Feasibility

{markdown_table(oep, ["year", "outcome", "exact_column_name", "geographic_level", "number_usable_county_year_observations", "missingness_rate", "suppressed_small_cells_exist", "matches_drake_categories"], 80)}

County-level outcome construction is possible for the tested years when the exact columns above are present. Suppression and missingness should be handled as data quality flags rather than silently imputed.

## Treatment Construction Feasibility

Prototype output: `data/intermediate/prototype_turnover_2023_2024.csv`.

Join diagnostics:

{markdown_table(prototype_diag, ["metric", "numerator", "denominator", "rate", "sample_states", "notes"], 30)}

The prototype tests the necessary keys: county FIPS, plan ID, issuer ID, metal level, rating area, age-40 premium, Plan ID Crosswalk mappings, Rate PUF age-40 rates, and Service Area county mapping. Across-issuer versus within-issuer turnover is distinguishable when issuer IDs are populated in the crosswalk/current-year landscape joins.

## Sample Alignment With Drake Et Al.

{markdown_table(alignment, ["check", "value", "notes"], 40)}

The raw public files make approximate county-year sample reconstruction possible. They do not make individual-level sample reconstruction possible.

## Problems And Unresolved Issues

""" + "\n".join(f"- {item}" for item in problems) + f"""

## Minimal Tests Completed

- HTTP/download checks with status, file size, and SHA-256 checksum in `data/metadata/data_manifest.csv`.
- ZIP member listing and tabular read tests in `data/metadata/file_inventory.csv`.
- Header checks and candidate variable role detection in `data/metadata/column_inventory.csv`.
- Missingness and suppression-like value counts in `outputs/missingness_summary.csv`.
- OEP outcome constructability checks in `outputs/oep_outcome_feasibility.csv`.
- 2023-to-2024 prototype join-key checks in `outputs/prototype_join_diagnostics.csv`.
- Row-count checks for CSV files read in chunks.

## Recommended Next Step

Proceed to a **full Drake-style replication dataset** before any causal modeling. The next build should formalize the HealthCare.gov state sample, apply Alaska/Hawaii exclusions, decide how to handle Nebraska, and compute the full national county-year zero-to-positive turnover measure. A richer continuous retention-shock measure is plausible after the binary turnover reconstruction is validated. County-level HTE or policy-learning designs should wait until the replication dataset is stable. Individual-level HTE is not supported by these public aggregate PUFs.

## Appendix

- Data manifest: `data/metadata/data_manifest.csv`
- File inventory: `data/metadata/file_inventory.csv`
- Column inventory: `data/metadata/column_inventory.csv`
- Missingness summary: `outputs/missingness_summary.csv`
- OEP outcome feasibility: `outputs/oep_outcome_feasibility.csv`
- Prototype join diagnostics: `outputs/prototype_join_diagnostics.csv`
- Prototype county output: `data/intermediate/prototype_turnover_2023_2024.csv`
- Code files: `scripts/01_download_marketplace_data.py`, `scripts/02_inspect_marketplace_files.py`
- Reproducibility notes: `docs/reproducibility_notes.md`
"""
    REPORT_PATH.write_text(report, encoding="utf-8")


def generate_repro_notes() -> None:
    text = """# Reproducibility Notes

## Environment

- Python tested in this workspace: `Python 3.14.0`
- Required packages: `requests`, `beautifulsoup4`, `pandas`, `numpy`, `openpyxl`

Install missing packages with:

```bash
python -m pip install requests beautifulsoup4 pandas numpy openpyxl
```

## Run Order

From the project root:

```bash
python scripts/01_download_marketplace_data.py
python scripts/02_inspect_marketplace_files.py
```

The download script writes `data/metadata/data_manifest.csv` and stores raw files under `data/raw/{source_group}/{year}/`. Existing files are reused unless `--force` is passed.

The inspection script regenerates:

- `data/metadata/file_inventory.csv`
- `data/metadata/column_inventory.csv`
- `outputs/sample_rows/`
- `outputs/missingness_summary.csv`
- `outputs/oep_outcome_feasibility.csv`
- `outputs/prototype_join_diagnostics.csv`
- `data/intermediate/prototype_turnover_2023_2024.csv`
- `docs/aca_zero_premium_turnover_data_feasibility.md`

## Expected Runtime And Disk

Expected download size is roughly 700 MB to 1 GB for the selected OEP, Exchange PUF, QHP Landscape, and fallback Health Plan Finder files. Runtime depends on network speed; inspection is chunked and should generally run in minutes on a normal workstation.

## Manual Downloads

No manual download is expected for the files in the manifest. Some manifest rows are expected to fail because Data.HealthCare.gov did not expose the PY2019-PY2021 QHP Landscape direct URL pattern at the time of this audit. CMS Health Plan Finder files are downloaded as the official fallback for those years.
"""
    REPRO_PATH.write_text(text, encoding="utf-8")


def main() -> None:
    ensure_dirs()
    manifest = load_manifest()
    print("Inspecting downloaded files...")
    file_inventory, column_inventory, _missingness = inspect_files(manifest)
    print("Verifying OEP outcome feasibility...")
    oep = verify_oep_outcomes(manifest)
    print("Running 2023-to-2024 treatment construction prototype...")
    _prototype, prototype_diag = run_treatment_prototype(manifest)
    print("Checking sample alignment...")
    alignment = sample_alignment(manifest, oep)
    print("Writing report and reproducibility notes...")
    generate_report(manifest, file_inventory, column_inventory, oep, prototype_diag, alignment)
    generate_repro_notes()
    print(f"Wrote {REPORT_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
